import os
import time
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from playwright.sync_api import sync_playwright

import sys
def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Setup caminhos
PROJECT_DIR = get_app_dir()
LIVROS_DIR = os.path.join(PROJECT_DIR, "livros")
USER_DATA_DIR = os.path.join(PROJECT_DIR, "chrome_profile")
SHEET_CONSOLIDACAO = "Consolidação"

# Gera o nome do arquivo de log unico para esta execucao
TIMESTAMP_EXECUCAO = datetime.now().strftime("%Y%m%d-%H%M%S")
LOG_FILE = os.path.join(LIVROS_DIR, f"log_execucao_{TIMESTAMP_EXECUCAO}.txt")

def translate_error(e):
    err_msg = str(e).lower()
    if "timeout" in err_msg:
        return "Tempo de espera esgotado (o site demorou a responder)"
    if "is not a function" in err_msg or "is not defined" in err_msg:
        return "Erro interno no script do portal da prefeitura"
    if "target closed" in err_msg:
        return "O Chrome foi fechado manualmente durante o processo"
    if "network" in err_msg:
        return "Erro de conexão com a internet ou portal fora do ar"
    if "selector" in err_msg or "not found" in err_msg or "visible" in err_msg:
        return "O robô não encontrou um elemento na tela (possível mudança no portal)"
    
    # Limpa excesso de detalhes tecnicos do Playwright
    clean_msg = str(e).split("=================")[0].strip()
    return f"Erro inesperado: {clean_msg}"

def log_exec(cnpj, nome, comp, status, detalhes, index=None, total=None, csv_path=None):
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    idx_str = f"[{index}/{total}] " if index and total else ""
    
    # Traduz detalhes se for erro
    detalhes_finais = translate_error(detalhes) if status == "ERRO" else detalhes
    
    linha_txt = f"[{agora}] {idx_str}CNPJ: {cnpj} | Nome: {nome} | Comp: {comp} | Status: {status} | Info: {detalhes_finais}\n"
    
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(linha_txt)
        
    if csv_path:
        # Se o CSV nao existir, cria cabecalho
        try:
            if not os.path.exists(csv_path):
                with open(csv_path, "w", encoding="utf-8-sig") as f:
                    f.write("Status;Index;Total;CNPJ;Representado;Competencia;Data_Hora;Detalhes\n")
            
            # Escapa ponto e virgula pra nao quebrar o CSV
            det_clean = str(detalhes_finais).replace(";", ",")
            linha_csv = f"{status};{index};{total};{cnpj};{nome};{comp};{agora};{det_clean}\n"
            with open(csv_path, "a", encoding="utf-8-sig") as f:
                f.write(linha_csv)
        except:
            pass # Nao trava o bot por erro de log

for d in [LIVROS_DIR, USER_DATA_DIR]:
    if not os.path.exists(d):
        os.makedirs(d)

URL_LOGIN = "https://nfse.campogrande.ms.gov.br/notafiscal/paginas/portal/index.html#/login"



def read_targets(config_file_path):
    if not os.path.exists(config_file_path):
        print(f"Arquivo {config_file_path} nao existe!")
        return []
        
    try:
        # Lê o excel sem assumir cabeçalho fixo primeiro
        df_raw = pd.read_excel(config_file_path, header=None)
        
        # Procura a linha que contém 'CNPJ' para ser o header de verdade
        header_idx = 0
        for i in range(min(10, len(df_raw))):
            row_vals = [str(v).lower() for v in df_raw.iloc[i].values]
            if any('cnpj' in v for v in row_vals):
                header_idx = i
                break
        
        # Re-lê com o header correto ou ajusta o DF
        df = pd.read_excel(config_file_path, header=header_idx)
        df = df.dropna(how='all') # Remove linhas totalmente vazias
            
        targets = []
        
        # Mapeamento dinâmico de colunas para suportar variacoes
        col_nome = None
        col_cnpj = None
        
        # 1. Tenta por nomes conhecidos
        possiveis_nomes = ['Razão Social', 'Nome', 'Nome_Empresa']
        possiveis_cnpjs = ['CNPJ', 'Cnpj', 'CPF/CNPJ']
        
        for col in df.columns:
            if any(p.lower() in str(col).lower() for p in possiveis_nomes):
                col_nome = col
            if any(p.lower() in str(col).lower() for p in possiveis_cnpjs):
                col_cnpj = col
        
        # 2. Se não achou por nome, tenta por índice (frequentemente 0 e 1 ou 1 e 2)
        if (col_nome is None or col_cnpj is None):
            if len(df.columns) >= 2:
                # Caso comum: Col 0 ou 1 é Nome, Col 1 ou 2 é CNPJ
                # No modelo gerado: 0=Razão, 1=CNPJ
                # No modelo prefeitura: 1=Razão, 2=CNPJ
                if pd.notna(df.iloc[0, 0]) and len(str(df.iloc[0, 1])) >= 11:
                    col_nome, col_cnpj = df.columns[0], df.columns[1]
                elif len(df.columns) >= 3:
                    col_nome, col_cnpj = df.columns[1], df.columns[2]
            else:
                print("Arquivo Excel com colunas insuficientes.")
                return []

        for _, row in df.iterrows():
            nome = str(row[col_nome]).strip() if pd.notna(row[col_nome]) else ""
            cnpj_raw = str(row[col_cnpj]).strip() if pd.notna(row[col_cnpj]) else ""
            
            # Limpar mascara do CNPJ
            cnpj_clean = "".join(filter(str.isdigit, cnpj_raw))
            
            if len(cnpj_clean) >= 11: # Suporta CPF ou CNPJ
                targets.append({
                    'Nome_Empresa': nome,
                    'CNPJ': cnpj_clean
                })
                
        print(f"Total de {len(targets)} empresas carregadas do arquivo.")
        return targets
    except Exception as e:
        print(f"Erro ao ler arquivo de alvos: {e}")
        return []

def process_livro_fiscal(filepath, target, competencia):
    print(f"Processando planilha: {filepath}")
    
    if filepath.lower().endswith('.pdf'):
        print(f"Erro: O arquivo baixado é um PDF ({filepath}). O menu Saída falhou em selecionar Excel. Pulando processamento Pandas...")
        return {"status": "ERRO_PDF"}
        
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side
        
        # O excel baixa com algumas linhas de cabecalho
        df = pd.read_excel(filepath, header=5)
        
        # Filtros por Situacao
        df_ativas = df[df['SITUACAO'].astype(str).str.lower() == 'ativa'].copy()
        df_cancel = df[df['SITUACAO'].astype(str).str.lower() == 'cancelada'].copy()
        
        qtd_ativas = len(df_ativas)
        qtd_cancel = len(df_cancel)
        
        # Keywords para soma de valores financeiros
        kv_finance = ['VALOR', 'RETIDO', 'ISS', 'BASE', 'CALCULO', 'DEDU', 'PIS', 'COFINS', 'CSLL', 'INSS', 'IR', 'DESCONTO', 'LIQUIDO']

        # Salva a coluna RETIDO (texto) e identifica a coluna VALOR ISS ANTES da conversão numérica
        col_retido = None
        col_iss_val = None
        retido_series_salva = None
        for col in df_ativas.columns:
            nome_col_upper = str(col).strip().upper()
            if nome_col_upper == 'RETIDO':
                col_retido = col
                retido_series_salva = df_ativas[col].astype(str).str.strip().str.lower().copy()
            if 'VALOR ISS' in nome_col_upper:
                col_iss_val = col

        # Helper para garantir tipagem numerica em todas colunas de valor
        def clean_curr(series):
            if series.dtype == 'object':
                return pd.to_numeric(series.astype(str).str.replace(r'[R$\s]', '', regex=True).str.replace('.', '', regex=False).str.replace(',', '.'), errors='coerce').fillna(0)
            return pd.to_numeric(series, errors='coerce').fillna(0)
            
        for col in df_ativas.columns:
            nome_col_upper = str(col).strip().upper()
            if any(x in nome_col_upper for x in kv_finance):
                df_ativas[col] = clean_curr(df_ativas[col])
                
        # Calcula totais por coluna
        totais = {}
        for col in df_ativas.columns:
            nome_col = str(col).strip().upper()
            if any(x in nome_col for x in kv_finance):
                totais[nome_col] = df_ativas[col].sum()

        # Calcula ISS separado por tipo de retenção usando a série de texto salva
        iss_nao_retido = 0
        iss_retido_val = 0
        if retido_series_salva is not None and col_iss_val is not None:
            mask_retido = retido_series_salva == 'retido'
            mask_nao_retido = retido_series_salva.str.contains('não retido|nao retido', na=False)
            iss_retido_val = df_ativas.loc[mask_retido, col_iss_val].sum()
            iss_nao_retido = df_ativas.loc[mask_nao_retido, col_iss_val].sum()

        try:
            # Escreve de volta no Excel usando openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            
            # Procura a ultima linha baseada no max_row
            row_total = ws.max_row + 2
            
            # Formata a legenda do total
            txt_label = f"TOTAIS (NOTAS ATIVAS: {qtd_ativas} | CANCELADAS: {qtd_cancel}):"
            cell_label = ws.cell(row=row_total, column=1, value=txt_label)
            cell_label.font = Font(bold=True)
            
            # Limpa a celula da coluna D no total (conforme solicitado pelo usuário para não poluir)
            ws.cell(row=row_total, column=4, value=None)

            for col_idx, col_name in enumerate(df.columns, start=1):
                nome_limpo = str(col_name).strip().upper()
                if nome_limpo in totais:
                    # Garantir que a coluna 4 não receba valor se coincidir com keyword
                    if col_idx == 4:
                        ws.cell(row=row_total, column=col_idx, value=None)
                        continue
                        
                    cell = ws.cell(row=row_total, column=col_idx, value=totais[nome_limpo])
                    cell.font = Font(bold=True)
                    cell.number_format = '#,##0.00'
                    
            # Escreve linhas de ISS por tipo de retenção abaixo do total
            if col_retido is not None and col_iss_val is not None:
                col_idx_retido = list(df.columns).index(col_retido) + 1 if col_retido in df.columns else None
                col_idx_iss = list(df.columns).index(col_iss_val) + 1 if col_iss_val in df.columns else None

                if col_idx_retido and col_idx_iss:
                    row_nr = row_total + 1
                    ws.cell(row=row_nr, column=col_idx_retido, value="Não Retido").font = Font(bold=True)
                    cell_nr = ws.cell(row=row_nr, column=col_idx_iss, value=iss_nao_retido)
                    cell_nr.font = Font(bold=True)
                    cell_nr.number_format = '#,##0.00'

                    row_r = row_total + 2
                    ws.cell(row=row_r, column=col_idx_retido, value="Retido").font = Font(bold=True)
                    cell_r = ws.cell(row=row_r, column=col_idx_iss, value=iss_retido_val)
                    cell_r.font = Font(bold=True)
                    cell_r.number_format = '#,##0.00'

            wb.save(filepath)
            
            # Retorna info para o resumo consolidado
            return {
                "status": "SUCESSO",
                "qtd_ativas": qtd_ativas,
                "qtd_cancel": qtd_cancel,
                "totais": totais,
                "iss_retido": iss_retido_val,
                "iss_nao_retido": iss_nao_retido
            }
        except Exception as e_xl:
            print(f"Aviso: Não foi possível adicionar totais com openpyxl (Arquivo provavel HTML?): {e_xl}")
            return {"status": "SUCESSO_PARCIAL", "qtd_ativas": qtd_ativas, "qtd_cancel": qtd_cancel, "totais": totais, "iss_retido": iss_retido_val, "iss_nao_retido": iss_nao_retido}

    except Exception as e:
        print(f"Erro ao processar totais localmente em {filepath}: {e}")
        return {"status": "ERRO_PANDAS", "msg": str(e)}

def run_bot(competencia_mes_ano, config_file_path, wait_for_input=False, output_dir=None, log_callback=None):
    targets = read_targets(config_file_path)
    if not targets:
        return
        

    # Setup do Log CSV dinâmico na pasta de destino
    out_dir = output_dir if output_dir else LIVROS_DIR
    errors_dir = os.path.join(out_dir, "erros")
    if not os.path.exists(errors_dir):
        os.makedirs(errors_dir)
        
    timestamp_log = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_log_path = os.path.join(out_dir, f"LogRoboNFSE-{timestamp_log}.csv")
    total_targets = len(targets)
    
    # Lista para consolidado final (Resumo)
    resumo_execucao = []

    print("Iniciando Playwright (Conectando ao Chrome aberto)...")
    with sync_playwright() as p:
        try:
            browser = p.chromium.connect_over_cdp('http://localhost:9222')
            default_context = browser.contexts[0]
            page = default_context.pages[0] if default_context.pages else default_context.new_page()
            
            # Só navega para o login se não estiver já no domínio da prefeitura
            if "nfse.campogrande.ms.gov.br" not in page.url:
                print(f"Navegando para a página inicial...")
                page.goto(URL_LOGIN, timeout=60000)
            else:
                print(f"Já estamos no portal (URL: {page.url}), continuando de onde parou...")
            
        except Exception as e:
            print(f"ERRO: Não foi possível conectar ao Chrome. Certifique-se de abri-lo pelo atalho de depuração.\nDetalhes: {e}")
            return
            
        if wait_for_input:
            input("Pressione ENTER SOMENTE quando você estiver na tela de 'Pesquisar CNPJ' após o Certificado Digital...")
        # VALIDAÇÃO DE SEGURANÇA: Garantir que o certificado/login é o da CONTILI
        try:
            page.wait_for_load_state('networkidle', timeout=5000)
            user_info = page.evaluate('() => document.body.innerText')
            if "CONTILI " not in str(user_info).upper() and "CONTILI\n" not in str(user_info).upper() and "CONTILI\t" not in str(user_info).upper():
                # Tratamento para garantir que pelo menos a palavra contili esteja visível nas identificações da prefeitura
                if "CONTILI" not in str(user_info).upper():
                    raise ValueError("ATENÇÃO: Certificado Incorreto! O robô detectou que você não está logado com a CONTILI. Cancele e logue novamente.")
            print("Validação de Segurança OK: Acesso CONTILI confirmado.")
            if log_callback: log_callback("✅ Acesso CONTILI confirmado. Iniciando loop...")
        except ValueError as ve:
            if log_callback: log_callback(f"❌ ERRO DE SEGURANÇA: {str(ve)}")
            raise ve # Dispara erro pro app_gui mostrar popup e fechar o botzinho
        except Exception as e:
            print(f"Aviso de Validação: Não foi possível raspar o nome do certificado na tela. Prosseguindo... ({e})")
            
        for target in targets:
            idx = targets.index(target) + 1
            cnpj = target['CNPJ'].strip()
            nome_empresa = target.get('Nome_Empresa', '')
            
            print(f">>> Iniciando target [{idx}/{total_targets}]: {cnpj}")
            if log_callback: log_callback(f"[{idx}/{total_targets}] Processando: {cnpj} - {nome_empresa}")
            

            
            
            # Funcao auxiliar para clicar usando javascript puro ignorando opacidade/visibilidade do JSF
            def js_click_text(selector, text):
                page.evaluate(f'''() => {{
                    Array.from(document.querySelectorAll("{selector}")).forEach(el => {{
                        if((el.innerText || el.textContent || "").includes("{text}") || (el.title || "").includes("{text}")) {{
                            el.click();
                        }}
                    }});
                }}''')
            
            try:
                # Voltar para home / pesquisa clicando no Menu "Seleciona Cadastro"
                # Abre o sidebar se estiver fechado (o botao Menu tem uma div com texto Menu e class hide-menu ou logo-menu)
                # Abre o sidebar se estiver fechado (o botao Menu tem uma div com texto Menu e class hide-menu ou logo-menu)
                try:
                    menu_btn = page.locator('.header-link.hide-menu').first
                    if menu_btn.is_visible():
                        menu_btn.click(timeout=1000)
                        time.sleep(1)
                except:
                    pass
                
                # Clica em 'Seleciona Cadastro' no menu lateral
                # Clica em 'Seleciona Cadastro' no menu lateral
                js_click_text('span.nav-label', 'Seleciona Cadastro')
                time.sleep(2)
            except Exception as e:
                print(f"Erro ao clicar em Seleciona Cadastro no menu: {e}")
                
            try:
                # 1. Informar o CNPJ e pesquisar
                page.fill('input[id*="idCpfCnpj"]', cnpj)
                js_click_text('a', 'Pesquisar')
                time.sleep(3)
                
                # Clicar na linha da tabela onde ta o CNPJ/check (Como o CNPJ vem formatado com mascara na web, clicamos no primeiro result)
                has_selecionar = page.evaluate('''() => {
                    let links = Array.from(document.querySelectorAll("a"));
                    return links.some(el => (el.textContent || "").includes("Selecionar") || (el.title || "").includes("Selecionar"));
                }''')
                
                if not has_selecionar:
                    print(f"CNPJ {cnpj} não listado nas procurações. Pulando pros próximos...")
                    if log_callback: log_callback(f"⚠️ {cnpj} não encontrado nas procurações.")
                    log_exec(cnpj, nome_empresa, competencia_mes_ano, "IGNORADO", "CNPJ não encontrado na tabela de procurações após pesquisa", index=idx, total=total_targets, csv_path=csv_log_path)
                    continue

                # Delay visual para o usuario acompanhar a busca
                time.sleep(2)

                # Usa clique nativo do Playwright para garantir os eventos do Primefaces (JSF) que trocam a empresa na sessão
                btn_selecionar = page.locator('a[title="Selecionar"], a:has-text("Selecionar")').first
                if btn_selecionar.is_visible():
                    btn_selecionar.click(force=True)
                else:
                    js_click_text('a', 'Selecionar')
                time.sleep(4)
                
                try: # Se houver accordion ou submenu lateral "Gerenciar NFSE" -> "Livro Fiscal"
                    js_click_text('span.nav-label', 'Gerenciar NFSe')
                    time.sleep(1)
                        
                    # Clica especificamente em Livro Fiscal no menu
                    js_click_text('span.nav-label', 'Livro Fiscal')
                    time.sleep(2)
                except Exception as e:
                    print(f"Nota: Erro ao clicar no menu Livro Fiscal: {translate_error(e)}")
                    page.screenshot(path=os.path.join(errors_dir, f"erro_menu_{cnpj}_{timestamp_log}.png"))
                    time.sleep(2)
                
                arquivos_baixados = {'ptd': None, 'tmd': None}
                # NOVO LOOP DE MÚLTIPLOS AGRUPAMENTOS: Prestados e Tomados
                for tipo_agrupamento, tipo_sigla in [('Serviços Prestados', 'ptd'), ('Serviços Tomados', 'tmd')]:
                    if log_callback: log_callback(f"Extraindo {tipo_agrupamento}...")
                    try:
                        # 3. Informar o Agrupamento no form do Primefaces
                        try:
                            lbl_agrup = page.locator('label', has_text='Agrupamento').first
                            if lbl_agrup.is_visible():
                                lbl_agrup.locator('xpath=..').locator('.ui-selectonemenu-trigger').click(force=True)
                                time.sleep(1)
                                item_tipo = page.locator(f'li.ui-selectonemenu-item:has-text("{tipo_agrupamento}")').first
                                if item_tipo.is_visible():
                                    item_tipo.click(force=True)
                                else:
                                    js_click_text('li.ui-selectonemenu-item', tipo_agrupamento)
                                time.sleep(1)
                        except Exception as e:
                            print(f"Nota: não foi possível mudar 'Agrupamento' para {tipo_agrupamento}. Erro: {e}")

                        # 4. Informar Datas e Filtros no form
                        # Aguarda AJAX do PrimeFaces estabilizar após troca de Agrupamento
                        time.sleep(3)
                        data_raw = competencia_mes_ano.replace("/", "")
                        
                        # Preenche Data Inicio com retry (AJAX do Agrupamento pode limpar o campo)
                        for tentativa in range(3):
                            inp_ini = page.locator('input[id$="idStart_input"]').first
                            inp_ini.click()
                            inp_ini.clear()
                            inp_ini.type(data_raw, delay=80)
                            time.sleep(0.5)
                            valor_preenchido = inp_ini.input_value()
                            if valor_preenchido and len(valor_preenchido) >= 6:
                                break
                            print(f"Retry {tentativa+1}: Campo Início vazio após preencher. Tentando novamente...")
                            time.sleep(1)
                        
                        inp_end = page.locator('input[id$="idEnd_input"]').first
                        inp_end.click()
                        inp_end.clear()
                        inp_end.type(data_raw, delay=80)

                        
                        # Marcar Todos situacao e Exigibilidade (Super Primefaces Checkbox Hack)
                        page.evaluate('''() => {
                            // Clica na caixa estilizada visual do Primefaces caso não esteja checada
                            document.querySelectorAll('.ui-chkbox-box:not(.ui-state-active)').forEach(el => el.click());
                            
                            // Força o checkbox escondido internamente para 'true' por segurança
                            document.querySelectorAll('input[type="checkbox"]').forEach(el => {
                                if (!el.checked) {
                                    el.checked = true;
                                    el.dispatchEvent(new Event('change', { bubbles: true }));
                                }
                            });
                        }''')
                                
                        # Clica Pesquisar para carregar a tabela!
                        js_click_text('a', 'Pesquisar')
                        time.sleep(4)
                                
                        # 5. Tentar encontrar a Saída "Excel" antes de clicar no export / gerar
                        try:
                            # 1. Clicar na caixa de dropdown Saída via Playwright para garantir que o evento Focus/Click ocorra no DOM real
                            lbl_saida = page.locator('label', has_text='Saída').first
                            if lbl_saida.is_visible():
                                # Clica na seta apontadora (trigger) do select correspondente
                                lbl_saida.locator('xpath=..').locator('.ui-selectonemenu-trigger').click(force=True)
                                time.sleep(1)
                                
                                # 2. Clicar na opção Excel na lista renderizada no fim do HTML
                                item_excel = page.locator('li.ui-selectonemenu-item:has-text("Excel"), li.ui-selectonemenu-item:has-text("XLS")').first
                                if item_excel.is_visible():
                                    item_excel.click(force=True)
                                else:
                                    js_click_text('li.ui-selectonemenu-item', 'Excel')
                                time.sleep(1)
                        except Exception as e:
                            print(f"Nota: não foi possível mudar a 'Saída' para Excel pelo dropdown. Continuando... Erro: {e}")

                        # 6. Baixar Excel
                        print(f"Aguardando geracao do excel ou mensagem de erro (vazio) para {tipo_sigla}...")
                        
                        btn_excel = None
                        for selector in [
                            'a:has-text("Download")', 
                            'button:has-text("Download")', 
                            'button:has-text("Excel")', 
                            'a:has-text("XLS")', 
                            'button:has-text("Gerar")', 
                            'button:has-text("Imprimir")', 
                            'a.ui-commandlink:has(i.pi-file-excel)'
                        ]:
                            element = page.locator(selector).first
                            if element.is_visible():
                                btn_excel = element
                                break
                                
                        if not btn_excel:
                            print(f"Botão de Excel não localizado para {tipo_sigla}. Capturando tela.")
                            page.screenshot(path=os.path.join(errors_dir, f"erro_btn_excel_{cnpj}_{tipo_sigla}_{timestamp_log}.png"))
                            btn_excel = page.locator('button[type="submit"]').first
                            
                        # Configura a captura do download em background
                        download_obj = [None]
                        def handle_download(d):
                            download_obj[0] = d
                            
                        page.on("download", handle_download)
                        btn_excel.click()
                        
                        # Loop de espera (Polling) para verificar se o download iniciou OU se deu erro de "Nenhum registro"
                        has_error = False
                        timeout_count = 0
                        while timeout_count < 120:
                            if download_obj[0]:
                                break
                                
                            # Verifica a mensagem de erro vermelha no painel informando que nao tem notas
                            if page.get_by_text("Nenhum registro", exact=False).first.is_visible():
                                has_error = True
                                break
                                    
                            page.wait_for_timeout(1000)
                            timeout_count += 1
                            
                        # Limpa o listener
                        page.remove_listener("download", handle_download)
                        
                        if has_error:
                            print(f"Empresa {cnpj} não possui notas {tipo_agrupamento} na competência {competencia_mes_ano}.")
                            if log_callback: log_callback(f"📂 {cnpj} não possui notas {tipo_sigla} no período (Zerado).")
                            log_exec(cnpj, nome_empresa, competencia_mes_ano, "VAZIO", f"Sem notas de {tipo_agrupamento}.", index=idx, total=total_targets, csv_path=csv_log_path)
                            resumo_execucao.append({
                                "Empresa": nome_empresa, "CNPJ": cnpj, "Tipo": tipo_agrupamento, 
                                "Competência": competencia_mes_ano, "Status": "Sem Movimento",
                                "Qtd Ativas": 0, "Qtd Cancel": 0, 
                                "Total Liquido": 0, "Total ISS": 0, "ISS Não Retido": 0, "ISS Retido": 0, "Base Calculo": 0,
                                "PIS": 0, "COFINS": 0, "CSLL": 0, "INSS": 0, "IR": 0, "Deducao": 0, "Desc": 0
                            })
                            continue 
                            
                        if not download_obj[0]:
                            log_exec(cnpj, nome_empresa, competencia_mes_ano, "ERRO_TIMEOUT", f"Timeout aguardando download {tipo_sigla}", index=idx, total=total_targets, csv_path=csv_log_path)
                            print(f"Timeout no {tipo_sigla} da empresa {cnpj}")
                            continue
                            
                        download = download_obj[0]
                        sugg_ext = download.suggested_filename.split('.')[-1]
                        if not sugg_ext: sugg_ext = 'xlsx'
                        
                        out_dir = output_dir if output_dir else LIVROS_DIR
                        
                        # Salva em arquivo temporário para consolidação posterior
                        temp_filepath = os.path.join(out_dir, f"temp_{tipo_sigla}_{cnpj}.xlsx")
                        if os.path.exists(temp_filepath):
                            os.remove(temp_filepath)
                            
                        download.save_as(temp_filepath)
                        arquivos_baixados[tipo_sigla] = temp_filepath
                        print(f"Download temporario salvo: {temp_filepath}")
                        
                        # 7. Processamento e Totais Ocultos
                        resultado_proc = process_livro_fiscal(temp_filepath, target, competencia_mes_ano)
                        
                        # Extrai totais para o resumo
                        v_total = 0; v_iss = 0; v_base = 0; v_pis = 0; v_cof = 0; v_ir = 0; v_csll = 0; v_inss = 0; v_deduc = 0; v_desc = 0
                        if "totais" in resultado_proc:
                            for k, v in resultado_proc["totais"].items():
                                if any(x in k for x in ['LIQUIDO', 'VALOR NF']): v_total += v
                                if 'ISS' in k: v_iss += v
                                if any(x in k for x in ['BASE', 'CALCULO']): v_base += v
                                if 'PIS' in k: v_pis += v
                                if 'COFINS' in k: v_cof += v
                                if 'IR' in k: v_ir += v
                                if 'CSLL' in k: v_csll += v
                                if 'INSS' in k: v_inss += v
                                if 'DEDU' in k: v_deduc += v
                                if 'DESC' in k: v_desc += v

                        v_iss_retido = resultado_proc.get("iss_retido", 0)
                        v_iss_nao_retido = resultado_proc.get("iss_nao_retido", 0)

                        resumo_execucao.append({
                            "Empresa": nome_empresa, "CNPJ": cnpj, "Tipo": tipo_agrupamento, 
                            "Competência": competencia_mes_ano, "Status": "Sucesso",
                            "Qtd Ativas": resultado_proc.get("qtd_ativas", 0),
                            "Qtd Cancel": resultado_proc.get("qtd_cancel", 0),
                            "Total Liquido": v_total, "Total ISS": v_iss, "ISS Não Retido": v_iss_nao_retido, "ISS Retido": v_iss_retido, "Base Calculo": v_base,
                            "PIS": v_pis, "COFINS": v_cof, "CSLL": v_csll, "INSS": v_inss, "IR": v_ir, "Deducao": v_deduc, "Desc": v_desc
                        })

                        if resultado_proc.get("status") == "SUCESSO":
                            log_exec(cnpj, nome_empresa, competencia_mes_ano, "SUCESSO", f"Arquivo {tipo_sigla.upper()} processado com Totais", index=idx, total=total_targets, csv_path=csv_log_path)
                        else:
                            log_exec(cnpj, nome_empresa, competencia_mes_ano, "SUCESSO_PARCIAL", f"Arquivo {tipo_sigla.upper()} baixado, falha totais", index=idx, total=total_targets, csv_path=csv_log_path)
                            
                    except Exception as loop_e:
                        print(f"Erro processando {tipo_sigla} de {cnpj}: {loop_e}")
                        log_exec(cnpj, nome_empresa, competencia_mes_ano, "ERRO_TIPO", f"Falha no tipo {tipo_sigla}: {loop_e}", index=idx, total=total_targets, csv_path=csv_log_path)
                
                # --- CONSOLIDAÇÃO DAS DUAS ABAS EM UM ÚNICO ARQUIVO ---
                nome_seguro = nome_empresa.replace('/', '-').replace('\\', '-').strip()
                if not nome_seguro:
                    nome_seguro = 'empresa'
                    
                out_dir = output_dir if output_dir else LIVROS_DIR
                base_filename = f"{competencia_mes_ano.replace('/','')}-NFSE-{nome_seguro}"
                final_filepath = os.path.join(out_dir, f"{base_filename}.xlsx")
                
                counter = 1
                while os.path.exists(final_filepath):
                    final_filepath = os.path.join(out_dir, f"{base_filename}({counter}).xlsx")
                    counter += 1
                    
                try:
                    import openpyxl
                    from copy import copy
                    
                    wb_final = openpyxl.Workbook()
                    wb_final.remove(wb_final.active)
                    
                    for key, title in [('ptd', 'Prestados'), ('tmd', 'Tomados')]:
                        fpath = arquivos_baixados[key]
                        if not fpath or not os.path.exists(fpath):
                            ws = wb_final.create_sheet(title)
                            ws.cell(1, 1, "Sem movimento para esta competência.")
                            continue
                            
                        wb_source = openpyxl.load_workbook(fpath)
                        ws_source = wb_source.active
                        ws_target = wb_final.create_sheet(title)
                        
                        for row in ws_source.iter_rows():
                            for cell in row:
                                new_cell = ws_target.cell(row=cell.row, column=cell.column, value=cell.value)
                                if cell.has_style:
                                    new_cell.font = copy(cell.font)
                                    new_cell.border = copy(cell.border)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.number_format = cell.number_format
                                    new_cell.protection = copy(cell.protection)
                                    new_cell.alignment = copy(cell.alignment)
                                    
                        for col_letter, col_dim in ws_source.column_dimensions.items():
                            ws_target.column_dimensions[col_letter].width = col_dim.width
                            
                        # Limpeza do temporário
                        wb_source.close()
                        os.remove(fpath)
                        
                    wb_final.save(final_filepath)
                    print(f"Arquivo consolidado gerado: {final_filepath}")
                    if log_callback: log_callback(f"💾 Arquivo consolidado salvo: {os.path.basename(final_filepath)}")
                except Exception as e_consol:
                    print(f"Erro ao consolidar abas para {cnpj}: {e_consol}")
                    if log_callback: log_callback(f"⚠️ Erro ao consolidar planilhas para {cnpj}")
            except Exception as e:
                msg_traduzida = translate_error(e)
                print(f"Erro ao processar Empresa {cnpj}: {msg_traduzida}")
                if log_callback: log_callback(f"❌ ERRO em {cnpj}: {msg_traduzida}")
                log_exec(cnpj, nome_empresa, competencia_mes_ano, "ERRO", e, index=idx, total=total_targets, csv_path=csv_log_path)
                page.screenshot(path=os.path.join(errors_dir, f"erro_final_{cnpj}_{timestamp_log}.png"))
                
        # --- FINAL: Gera o arquivo de Resumo Consolidado ---
        if resumo_execucao:
            try:
                df_resumo = pd.DataFrame(resumo_execucao)
                resumo_path = os.path.join(out_dir, f"Resumo_Consolidado_{competencia_mes_ano.replace('/','')}_{timestamp_log}.xlsx")
                df_resumo.to_excel(resumo_path, index=False)
                
                # --- Pós-Processamento Visual ---
                import openpyxl
                from openpyxl.styles import Border, Side
                wb_resumo = openpyxl.load_workbook(resumo_path)
                ws_resumo = wb_resumo.active
                
                thick = Side(border_style='thick', color="000000")
                
                # Formata valores e adiciona bordas grossas a cada 2 linhas (agrupamento de empresa)
                for row_idx in range(2, ws_resumo.max_row + 1):
                    # Formata colunas de valores (F em diante)
                    for col_idx in range(6, ws_resumo.max_column + 1):
                        ws_resumo.cell(row_idx, col_idx).number_format = '#,##0.00'
                        
                    # Se for linha par (Tomados), adiciona borda grossa embaixo
                    if row_idx % 2 == 1: # Cabeçalho é 1. Prestado é 2. Tomado é 3.
                        # Na verdade, a empresa ocupa 2 e 3. Borda embaixo da 3.
                        pass
                    
                    if (row_idx - 1) % 2 == 0: 
                        # row 3, 5, 7...
                        for c_idx in range(1, ws_resumo.max_column + 1):
                            ws_resumo.cell(row_idx, c_idx).border = Border(bottom=thick)

                wb_resumo.save(resumo_path)
                print(f"Arquivo de Resumo Consolidado gerado e formatado: {resumo_path}")
                if log_callback: log_callback(f"📊 Resumo Consolidado formatado em dist/livros")
            except Exception as eerr:
                print(f"Erro ao gerar resumo consolidado: {eerr}")

        print("Finalizado processamento de todas as empresas da lista!")
        if log_callback: log_callback("🏁 FIM DO PROCESSO: Todas as empresas foram verificadas.")
        browser.close()

if __name__ == "__main__":
    competencia = input("Digite a competência para buscar os livros (Formato MM/AAAA): ")
    # Em caso de rodar solto no terminal pro debug manual, fixa o padrao antigo
    old_config = os.path.join(PROJECT_DIR, "planejamento", "Relação de empresas.xlsx")
    run_bot(competencia, old_config)
